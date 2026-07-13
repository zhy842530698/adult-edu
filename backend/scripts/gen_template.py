"""Generate the Excel import template."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
HEADER_FONT = Font(bold=True)


REQUIRED = [
    "exam_code", "subject_code", "chapter_code", "knowledge_codes",
    "question_type", "stem", "option_a", "option_b", "option_c", "option_d",
    "option_e", "option_f", "option_g", "option_h", "answer",
    "analysis", "difficulty", "score", "source_name", "source_year",
    "source_question_no", "license_type", "source_type", "real_exam_year",
    "external_ref", "assets", "tags",
]


def main(out_path: str | None = None) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "题库导入"
    for col_idx, name in enumerate(REQUIRED, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")
    # example row
    ws.append([
        "EN", "LISTENING", "LONG_DIALOGUE", "MAIN_IDEA",
        "SINGLE_CHOICE",
        "What is the main topic of the conversation?",
        "Booking a hotel", "Ordering food", "Asking for directions", "Talking about weather",
        "", "", "", "",
        "A",
        "对话中双方围绕预订酒店展开",
        3, 2.0, "CET4-2023-06", 2023, "1-A", "platform-original",
        "REAL_EXAM", 2023,
        "",
        "/static/audios/cet4-2023-06-q1.mp3",
        "sample,main_idea",
    ])
    # column widths
    for i, h in enumerate(REQUIRED, start=1):
        col_letter = ""
        n = i
        while n > 0:
            n, rem = divmod(n - 1, 26)
            col_letter = chr(ord("A") + rem) + col_letter
        ws.column_dimensions[col_letter].width = max(14, len(str(h)) + 2)

    notes = wb.create_sheet("字段说明")
    notes_rows = [
        ("exam_code", "必填；考试编码，如 EN/CET4"),
        ("subject_code", "必填；科目编码，如 LISTENING"),
        ("chapter_code", "选填；章节编码"),
        ("knowledge_codes", "选填；知识点编码，英文逗号分隔"),
        ("question_type", "必填；SINGLE_CHOICE 或 MULTIPLE_CHOICE"),
        ("stem", "必填；题干，允许受控 HTML"),
        ("option_a..h", "至少 2 项；选项内容"),
        ("answer", "必填；单选如 B；多选如 A,C"),
        ("analysis", "必填；解析"),
        ("difficulty", "必填；1-5"),
        ("score", "必填；正数"),
        ("source_name", "必填；来源或'平台原创'"),
        ("source_year", "选填"),
        ("source_question_no", "选填"),
        ("license_type", "必填；授权类型"),
        ("source_type", "选填；PLATFORM_ORIGINAL / REAL_EXAM / MOCK / COMPILATION，留空按 PLATFORM_ORIGINAL"),
        ("real_exam_year", "选填；仅 source_type=REAL_EXAM 时填，如 2020"),
        ("external_ref", "选填"),
        ("assets", "选填；图片/音频 URL，逗号分隔；按扩展名自动判断 IMAGE/AUDIO；可写 'IMAGE:url|AUDIO:url' 显式标注类型"),
        ("tags", "选填；英文逗号分隔"),
    ]
    for row in notes_rows:
        notes.append(row)
    notes.column_dimensions["A"].width = 24
    notes.column_dimensions["B"].width = 60

    out = out_path or str(Path(__file__).resolve().parent / "excel_import_template.xlsx")
    wb.save(out)
    return out


if __name__ == "__main__":
    p = main()
    print(f"[template] 生成：{p}")

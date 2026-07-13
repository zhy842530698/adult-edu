"""PDF -> Excel conversion.

Parses an exam-paper PDF into the question-import Excel template so admins can
upload it through the existing "批量导入" flow.

Supported formats (best-effort):
- 题号:    "1." / "1、" / "1．" / "(1)" / "（1）" / "第1题"
- 选项:    "A." / "A、" / "(A)" / "（A）" / "A：" / "A)"
- 答案:    "答案: A,B" / "【答案】A" / "参考答案: A" / "Answer: A"
- 解析:    "解析: ..." / "【解析】..." / "解析: ..."

For PDFs without detectable answers/解析, those cells are left blank — admins
fill them by hand in the generated Excel before importing.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Reuse the canonical column order from gen_template.py so the output can be
# fed straight into /admin/import-jobs.
try:  # scripts/ is on sys.path via app/scripts_runtime.py at runtime
    from gen_template import REQUIRED, HEADER_FILL, HEADER_FONT  # type: ignore
except Exception:  # pragma: no cover — running the file directly
    REQUIRED = [
        "exam_code", "subject_code", "chapter_code", "knowledge_codes",
        "question_type", "stem", "option_a", "option_b", "option_c", "option_d",
        "option_e", "option_f", "option_g", "option_h", "answer",
        "analysis", "difficulty", "score", "source_name", "source_year",
        "source_question_no", "license_type", "source_type", "real_exam_year",
        "external_ref", "assets", "tags",
    ]
    HEADER_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    HEADER_FONT = Font(bold=True)


OPTION_LETTERS = "ABCDEFGH"
OPTION_COLUMNS = [f"option_{c.lower()}" for c in OPTION_LETTERS]

# ---- regex catalog --------------------------------------------------------

# 题号 — must START a line so we don't false-positive on "Reading 1. ..."
# Tightened: a number alone on a line (followed by 题 / punctuation) is a question.
# A number followed by letters/words (e.g. "2023 CET-4 Listening") is NOT.
_QNUM_RE = re.compile(
    r"""^\s*
        (?:第\s*)?
        (?:
            \(\s*(\d+)\s*\)          |  # (1)
            （\s*(\d+)\s*）          |  # （1）
            (\d+)\s*[.、．]          |  # 1. / 1、 / 1．
            (\d+)\s*题                 # 1题 / 1 题
        )
    """,
    re.MULTILINE | re.VERBOSE,
)

# 资产 URL 提取 — http(s)://... 加可选的显式类型前缀
# 例: "AUDIO:https://cdn/x.mp3" / "IMAGE:https://cdn/x.png"
_ASSET_RE = re.compile(
    r"""
    (?:IMAGE|AUDIO|图片|音频|语音)\s*[:：]\s*(\S+\.(?:jpg|jpeg|png|gif|webp|mp3|m4a|wav|ogg|mp4))   |
    (https?://\S+\.(?:jpg|jpeg|png|gif|webp|mp3|m4a|wav|ogg|mp4))
    """,
    re.IGNORECASE | re.VERBOSE,
)

# 选项 — line-leading "A." / "(A)" / "A、".  Falls back to a bare letter on
# its own line (pdfplumber sometimes strips punctuation from option rows).
_OPT_RE = re.compile(
    r"""^\s*
        (?:\(\s*([A-H])\s*\)|（\s*([A-H])\s*）)
        \s*[.、．\)）:：]?
        |
        ^\s*([A-H])
        \s*[.、．\)）:：]
    """,
    re.MULTILINE | re.VERBOSE,
)

# Bare-letter option line ("A\nB\nC\nD") — used when the primary regex misses
# a question entirely because pdfplumber dropped the punctuation.
_BARE_OPT_LINE_RE = re.compile(r"^\s*([A-H])\s*$", re.MULTILINE)

# 答案 — anywhere on a line, allow Chinese / English / brackets / multiple letters
_ANS_RE = re.compile(
    r"""
    (?:【\s*答案\s*】|答\s*案\s*[::]|参\s*考\s*答\s*案\s*[::]|
       Answer\s*:|Ans\s*:|答案\s*[::]|参考答案)
    \s*([A-H](?:\s*[\s,，、]\s*[A-H])*)
    """,
    re.IGNORECASE | re.VERBOSE,
)

# Same markers as _ANS_RE, used for trimming option content.
_ANS_TRIM_RE = re.compile(
    r"(?:【\s*答案\s*】|答\s*案\s*[::]|参\s*考\s*答\s*案\s*[::]"
    r"|Answer\s*:|Ans\s*:|Answer\s+is)",
    re.IGNORECASE,
)

# 解析 — text after the marker, captured up to the next 题号 or end of text
_ANALYSIS_RE = re.compile(
    r"""
    (?:【\s*解\s*析\s*】|解\s*析\s*[::：]|分\s*析\s*[::：]|Analysis\s*:|解析)
    \s*(.+?)(?=
        \n\s*(?:第\s*)?(?:\(\s*\d+\s*\)|（\s*\d+\s*）|\b\d+\b)
        \s*(?:题|[.、．\)）\s])
        | \Z
    )
    """,
    re.IGNORECASE | re.VERBOSE | re.DOTALL,
)

# Trim pattern for the same markers (English + Chinese).
_ANALYSIS_TRIM_RE = re.compile(
    r"(?:【\s*解\s*析\s*】|解\s*析\s*[::：]|分\s*析\s*[::：]|Analysis\s*:|解析)",
    re.IGNORECASE,
)


# ---- data classes --------------------------------------------------------

@dataclass
class ParsedQuestion:
    number: int | None = None
    stem: str = ""
    options: dict[str, str] = field(default_factory=dict)  # A→content
    answer: str = ""        # "A" or "A,C"
    analysis: str = ""
    assets: list[dict] = field(default_factory=list)  # [{"asset_type": "IMAGE"|"AUDIO", "url": "..."}]


@dataclass
class ParseResult:
    questions: list[ParsedQuestion]
    warnings: list[str] = field(default_factory=list)
    pages: int = 0

    @property
    def ok_count(self) -> int:
        return sum(1 for q in self.questions if q.stem and len(q.options) >= 2)

    @property
    def missing_answer(self) -> int:
        return sum(1 for q in self.questions if q.stem and not q.answer)


# ---- core parsing ---------------------------------------------------------

def _read_pdf_text(file_bytes: bytes) -> tuple[str, int]:
    text_parts: list[str] = []
    pages = 0
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        pages = len(pdf.pages)
        for page in pdf.pages:
            t = page.extract_text() or ""
            if t:
                text_parts.append(t)
    return "\n".join(text_parts), pages


def _split_into_blocks(text: str) -> list[tuple[int, str]]:
    """Return [(question_number, block_text)] — each block is one question.

    The block includes the trailing 答案/解析 until the next 题号.
    """
    matches = list(_QNUM_RE.finditer(text))
    blocks: list[tuple[int, str]] = []
    for i, m in enumerate(matches):
        # recover the captured number (any of the four groups)
        num = next((g for g in m.groups() if g), None)
        try:
            n = int(num) if num else None
        except ValueError:
            n = None
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        block = text[start:end]
        blocks.append((n, block))
    return blocks


def _parse_block(num: int, block: str) -> ParsedQuestion:
    q = ParsedQuestion(number=num)

    # 1) options — anything that matches a line-leading option marker
    opt_matches = list(_OPT_RE.finditer(block))
    if opt_matches:
        first_opt_start = opt_matches[0].start()
        # find the start of the FIRST option LINE (the option marker might
        # not be at column 0 — but in our regex it is, so start = match.start())
        q.stem = block[:first_opt_start].strip()
        q.stem = _QNUM_RE.sub("", q.stem, count=1).strip()

        for i, om in enumerate(opt_matches):
            letter = next((g for g in om.groups() if g), "").upper()
            content_start = om.end()
            content_end = opt_matches[i + 1].start() if i + 1 < len(opt_matches) else len(block)
            content = block[content_start:content_end].strip()
            content = _ANS_TRIM_RE.split(content, maxsplit=1)[0].strip()
            content = _ANALYSIS_TRIM_RE.split(content, maxsplit=1)[0].strip()
            q.options[letter] = content
    else:
        # ---- fallback: pdfplumber sometimes returns bare "A" / "B" lines
        # with no punctuation. Treat a run of 2-8 consecutive single-letter
        # lines as the option block.
        letters_in_block: list[str] = []
        for bm in _BARE_OPT_LINE_RE.finditer(block):
            letters_in_block.append(bm.group(1))
            if len(letters_in_block) >= 8:
                break
        if 2 <= len(letters_in_block) <= 8 and letters_in_block == list(OPTION_LETTERS[: len(letters_in_block)]):
            bare = list(_BARE_OPT_LINE_RE.finditer(block))
            # strip 题号 from the first line, before the first option
            q.stem = _QNUM_RE.sub("", block[: bare[0].start()], count=1).strip()
            for i, bm in enumerate(bare):
                if i + 1 < len(bare):
                    content = block[bm.end(): bare[i + 1].start()].strip()
                else:
                    content = block[bm.end():].strip()
                content = _ANS_TRIM_RE.split(content, maxsplit=1)[0].strip()
                content = _ANALYSIS_TRIM_RE.split(content, maxsplit=1)[0].strip()
                q.options[bm.group(1)] = content

        if not q.options:
            # truly no options → entire block is stem
            q.stem = _QNUM_RE.sub("", block, count=1).strip()
            q.stem = _ANS_TRIM_RE.split(q.stem, maxsplit=1)[0].strip()
            q.stem = _ANALYSIS_TRIM_RE.split(q.stem, maxsplit=1)[0].strip()

    # 2) answer
    am = _ANS_RE.search(block)
    if am:
        raw = am.group(1)
        letters = [c for c in re.split(r"[\s,，、]+", raw.upper()) if c in OPTION_LETTERS]
        q.answer = ",".join(sorted(set(letters)))

    # 3) analysis
    anm = _ANALYSIS_RE.search(block)
    if anm:
        q.analysis = " ".join(anm.group(1).split())

    # 4) assets — look for inline image/audio URLs in the block
    for m in _ASSET_RE.finditer(block):
        url = m.group(1) or m.group(2)
        if not url:
            continue
        explicit_type = (m.group(0).split(":")[0].split("：")[0] or "").upper()
        if explicit_type in ("IMAGE", "图片"):
            asset_type = "IMAGE"
        elif explicit_type in ("AUDIO", "音频", "语音"):
            asset_type = "AUDIO"
        else:
            lower = url.lower()
            if any(lower.endswith(ext) for ext in (".jpg", ".jpeg", ".png", ".gif", ".webp")):
                asset_type = "IMAGE"
            elif any(lower.endswith(ext) for ext in (".mp3", ".m4a", ".wav", ".ogg", ".mp4")):
                asset_type = "AUDIO"
            else:
                asset_type = "IMAGE"
        q.assets.append({"asset_type": asset_type, "url": url.rstrip(".,;)")})

    return q


def parse_pdf(file_bytes: bytes) -> ParseResult:
    text, pages = _read_pdf_text(file_bytes)
    if not text.strip():
        return ParseResult(questions=[], warnings=["PDF 中未提取到任何文字（可能是扫描件或图片型 PDF）"], pages=pages)

    blocks = _split_into_blocks(text)
    if not blocks:
        return ParseResult(questions=[], warnings=["未识别到任何题号（题干前请加 '1.' '2.' 等题号）"], pages=pages)

    questions: list[ParsedQuestion] = []
    no_answer_warn = 0
    for num, block in blocks:
        q = _parse_block(num, block)
        if not q.stem:
            continue  # skip noise blocks that look like 题号 but aren't real
        if not q.answer:
            no_answer_warn += 1
        questions.append(q)

    warnings: list[str] = []
    if no_answer_warn:
        warnings.append(f"{no_answer_warn} 道题未在 PDF 中识别到答案，需在生成 Excel 中手动补齐")
    return ParseResult(questions=questions, warnings=warnings, pages=pages)


# ---- Excel writer ---------------------------------------------------------

def _build_excel(
    questions: Iterable[ParsedQuestion],
    *,
    exam_code: str,
    subject_code: str,
    chapter_code: str = "",
    knowledge_codes: str = "",
    source_name: str = "",
    source_year: int | None,
    source_type: str = "PLATFORM_ORIGINAL",
    real_exam_year: int | None = None,
    license_type: str = "platform-original",
    difficulty: int = 3,
    score: float = 2.0,
    tags: str = "",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "PDF 解析结果"

    # header
    for col_idx, name in enumerate(REQUIRED, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")
    for i, h in enumerate(REQUIRED, start=1):
        col_letter = ""
        n = i
        while n > 0:
            n, rem = divmod(n - 1, 26)
            col_letter = chr(ord("A") + rem) + col_letter
        ws.column_dimensions[col_letter].width = max(14, len(str(h)) + 2)

    row_no = 2
    for q in questions:
        values = {col: "" for col in REQUIRED}

        values["exam_code"] = exam_code
        values["subject_code"] = subject_code
        values["chapter_code"] = chapter_code
        values["knowledge_codes"] = knowledge_codes

        # 单/多选判断：有 "," 的答案 = 多选题
        if "," in q.answer:
            values["question_type"] = "MULTIPLE_CHOICE"
        else:
            values["question_type"] = "SINGLE_CHOICE"

        values["stem"] = q.stem
        for letter in OPTION_LETTERS:
            col = f"option_{letter.lower()}"
            values[col] = q.options.get(letter, "")

        values["answer"] = q.answer
        values["analysis"] = q.analysis or ""
        values["difficulty"] = difficulty
        values["score"] = score

        values["source_name"] = source_name or f"PDF-导入-{source_year or ''}".rstrip("-")
        values["source_year"] = source_year or ""
        values["source_question_no"] = str(q.number) if q.number is not None else ""
        values["license_type"] = license_type
        values["source_type"] = source_type
        values["real_exam_year"] = real_exam_year or ""
        # assets — Excel cell value is comma-separated "[TYPE:]URL" entries
        asset_entries = []
        for a in getattr(q, "assets", []):
            url = a["url"]
            asset_entries.append(f"{a['asset_type']}:{url}" if not url.lower().startswith(("http://", "https://", "/static/")) else url)
        values["assets"] = ",".join(asset_entries)
        values["tags"] = tags

        for col_idx, name in enumerate(REQUIRED, start=1):
            ws.cell(row=row_no, column=col_idx, value=values[name])
        row_no += 1

    notes = wb.create_sheet("字段说明")
    notes_rows = [
        ("来源", "本文件由 PDF 解析工具自动生成"),
        ("必填提示", "answer / analysis 列若为空，请人工补齐后再走导入流程"),
    ]
    for r in notes_rows:
        notes.append(r)
    notes.column_dimensions["A"].width = 16
    notes.column_dimensions["B"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def convert_pdf_to_excel(
    file_bytes: bytes,
    *,
    exam_code: str,
    subject_code: str,
    source_year: int | None,
    is_real_exam: bool,
    remark: str = "",
    source_name: str = "",
    license_type: str = "platform-original",
    chapter_code: str = "",
    knowledge_codes: str = "",
    difficulty: int = 3,
    score: float = 2.0,
) -> tuple[bytes, ParseResult]:
    """Top-level helper used by the API: returns (xlsx_bytes, parse_summary)."""
    result = parse_pdf(file_bytes)

    source_type = "REAL_EXAM" if is_real_exam else "MOCK"
    real_exam_year = source_year if is_real_exam else None

    # 把表单的 remark 拼到 source_name，方便后续审计
    name = source_name or ""
    if remark and not name:
        name = remark
    elif remark:
        name = f"{name}-{remark}"

    xlsx = _build_excel(
        result.questions,
        exam_code=exam_code,
        subject_code=subject_code,
        chapter_code=chapter_code,
        knowledge_codes=knowledge_codes,
        source_name=name or (f"{exam_code}-{source_year}" if source_year else f"{exam_code}-PDF"),
        source_year=source_year,
        source_type=source_type,
        real_exam_year=real_exam_year,
        license_type=license_type,
        difficulty=difficulty,
        score=score,
        tags=remark,
    )
    return xlsx, result


if __name__ == "__main__":  # pragma: no cover — manual smoke test
    import sys

    if len(sys.argv) < 2:
        print("usage: pdf_to_excel.py <pdf> [exam_code] [subject_code] [year]")
        sys.exit(1)
    pdf_path = Path(sys.argv[1])
    exam = sys.argv[2] if len(sys.argv) > 2 else "EN"
    subj = sys.argv[3] if len(sys.argv) > 3 else "LISTENING"
    year = int(sys.argv[4]) if len(sys.argv) > 4 else 2023
    data = pdf_path.read_bytes()
    xlsx, res = convert_pdf_to_excel(
        data,
        exam_code=exam,
        subject_code=subj,
        source_year=year,
        is_real_exam=True,
        remark="smoke-test",
    )
    out = pdf_path.with_suffix(".xlsx")
    out.write_bytes(xlsx)
    print(f"[pdf] {res.pages} pages, {len(res.questions)} questions, ok={res.ok_count}, missing_answer={res.missing_answer}")
    for w in res.warnings:
        print(f"[warn] {w}")
    print(f"[xlsx] {out}")
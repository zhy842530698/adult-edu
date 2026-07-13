"""Excel import parser — per-row validation."""
from __future__ import annotations

import io
import json
import re
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook

from app.services.html_sanitizer import clean as clean_html


REQUIRED_COLUMNS = [
    "exam_code", "subject_code", "question_type", "stem",
    "option_a", "option_b", "answer", "analysis", "difficulty", "score",
    "source_name", "license_type",
]
# assets is OPTIONAL — adding it to REQUIRED would break older Excel files.
OPTIONAL_COLUMNS = ["assets"]
OPTION_COLUMNS = [f"option_{c.lower()}" for c in "ABCDEFGH"]

SOURCE_TYPE_VALUES = {"PLATFORM_ORIGINAL", "REAL_EXAM", "MOCK", "COMPILATION"}


_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}
_AUDIO_EXTS = {".mp3", ".m4a", ".wav", ".ogg", ".mp4", ".aac", ".flac"}


def _parse_assets(raw: str) -> list[dict]:
    """Parse the assets cell into [{asset_type, url}, ...].

    Accepted formats (comma-separated):
      https://cdn/x.jpg,https://cdn/y.mp3               # auto-detect by ext
      IMAGE:https://cdn/x.jpg|AUDIO:/static/z.mp3       # pipe also accepted
      图片:https://..., 音频:...                          # Chinese prefixes
    Empty / whitespace → [].
    """
    if not raw:
        return []
    out: list[dict] = []
    # allow either comma or pipe as separator
    parts = re.split(r"[,|]", raw)
    for p in parts:
        p = p.strip()
        if not p:
            continue
        url = p
        asset_type: str | None = None
        # explicit prefix?
        for prefix, kind in (("IMAGE:", "IMAGE"), ("AUDIO:", "AUDIO"),
                             ("图片:", "IMAGE"), ("音频:", "AUDIO"), ("语音:", "AUDIO")):
            if p.upper().startswith(prefix.upper()):
                asset_type = kind
                url = p[len(prefix):].strip()
                break
        if asset_type is None:
            ext = "." + url.rsplit(".", 1)[-1].lower() if "." in url.rsplit("/", 1)[-1] else ""
            if ext in _IMAGE_EXTS:
                asset_type = "IMAGE"
            elif ext in _AUDIO_EXTS:
                asset_type = "AUDIO"
            else:
                # unknown — skip but don't error; admins can fix in edit page
                continue
        out.append({"asset_type": asset_type, "url": url})
    return out


@dataclass
class RowResult:
    row_no: int
    status: str  # OK / WARN / ERROR
    payload: dict | None = None
    errors: list[str] = field(default_factory=list)


def _coerce_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_rows(file_bytes: bytes) -> list[RowResult]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(h or "").strip() for h in next(rows_iter)]
    except StopIteration:
        return []
    # Build header index
    idx = {h: i for i, h in enumerate(header)}
    missing = [c for c in REQUIRED_COLUMNS if c not in idx]
    if missing:
        return [RowResult(row_no=0, status="ERROR", errors=[f"缺少必填列: {', '.join(missing)}"])]

    results: list[RowResult] = []
    row_no = 1  # data row 1 is the first data row after header
    for raw in rows_iter:
        row_no += 1
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        try:
            cells = [_coerce_cell(c) for c in raw]
        except Exception as exc:
            results.append(RowResult(row_no=row_no, status="ERROR", errors=[f"行解析失败: {exc}"]))
            continue
        errors: list[str] = []
        payload: dict[str, Any] = {}

        def cell(col: str) -> str:
            i = idx.get(col)
            if i is None or i >= len(cells):
                return ""
            return cells[i]

        # exam / subject / chapter / kp codes
        payload["exam_code"] = cell("exam_code")
        payload["subject_code"] = cell("subject_code")
        payload["chapter_code"] = cell("chapter_code") or None
        payload["knowledge_codes"] = [c.strip() for c in (cell("knowledge_codes") or "").split(",") if c.strip()]

        # question type
        qt = cell("question_type").upper()
        if qt not in ("SINGLE_CHOICE", "MULTIPLE_CHOICE"):
            errors.append(f"question_type 必须是 SINGLE_CHOICE 或 MULTIPLE_CHOICE（当前：{qt}）")
        payload["question_type"] = qt

        # options
        options = []
        for i, col in enumerate(OPTION_COLUMNS, start=0):
            v = cell(col)
            if v:
                options.append({"option_code": chr(ord("A") + i), "content": v})
        payload["options"] = options

        if len(options) < 2:
            errors.append("至少需要 2 个选项")
        if qt == "MULTIPLE_CHOICE" and len(options) < 3:
            errors.append("多选题至少需要 3 个选项")

        # stem / analysis
        payload["stem"] = cell("stem")
        payload["analysis"] = cell("analysis")
        if not payload["stem"]:
            errors.append("题干 stem 不能为空")
        if not payload["analysis"]:
            errors.append("解析 analysis 不能为空")

        # answer
        ans_raw = cell("answer")
        answer_codes = [c.strip().upper() for c in ans_raw.split(",") if c.strip()]
        valid_codes = [o["option_code"] for o in options]
        unknown = [c for c in answer_codes if c not in valid_codes]
        if unknown:
            errors.append(f"答案 {unknown} 不在现有选项中")
        if qt == "SINGLE_CHOICE":
            if len(answer_codes) != 1:
                errors.append("单选题答案必须恰好 1 个")
        elif qt == "MULTIPLE_CHOICE":
            if len(set(answer_codes)) < 2:
                errors.append("多选题答案至少 2 个")
        payload["correct_options"] = sorted(set(answer_codes))

        # difficulty / score
        try:
            d = int(cell("difficulty"))
            if not (1 <= d <= 5):
                errors.append("难度 difficulty 必须在 1-5 之间")
            payload["difficulty"] = d
        except ValueError:
            errors.append("难度 difficulty 必须是整数")
            payload["difficulty"] = None
        try:
            s = float(cell("score"))
            if s <= 0:
                errors.append("分值 score 必须为正数")
            payload["score"] = s
        except ValueError:
            errors.append("分值 score 必须是数字")
            payload["score"] = None

        # source
        payload["source_name"] = cell("source_name")
        payload["source_year"] = cell("source_year") or None
        payload["source_question_no"] = cell("source_question_no") or None
        payload["license_type"] = cell("license_type")
        payload["external_ref"] = cell("external_ref") or None
        payload["assets"] = _parse_assets(cell("assets"))
        payload["tags"] = [t.strip() for t in (cell("tags") or "").split(",") if t.strip()]

        # structured source classification (column is optional — empty falls
        # back to PLATFORM_ORIGINAL)
        st = cell("source_type").upper()
        if st and st not in SOURCE_TYPE_VALUES:
            errors.append(
                f"source_type 必须是 {sorted(SOURCE_TYPE_VALUES)} 之一（当前：{st or '空'}）"
            )
        payload["source_type"] = st or "PLATFORM_ORIGINAL"
        rey = cell("real_exam_year")
        if rey:
            try:
                payload["real_exam_year"] = int(rey)
            except ValueError:
                errors.append("real_exam_year 必须是整数年份")
                payload["real_exam_year"] = None
        else:
            payload["real_exam_year"] = None
        if payload["source_type"] == "REAL_EXAM" and not payload["real_exam_year"]:
            errors.append("source_type=REAL_EXAM 必须填 real_exam_year")

        if not payload["source_name"]:
            errors.append("source_name 不能为空")
        if not payload["license_type"]:
            errors.append("license_type 不能为空")

        if errors:
            results.append(RowResult(row_no=row_no, status="ERROR", errors=errors))
        else:
            results.append(RowResult(row_no=row_no, status="OK", payload=payload))
    return results
